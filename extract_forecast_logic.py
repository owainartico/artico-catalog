import openpyxl
import json

def extract_fytm_percentages(file_path):
    """Extract monthly sales percentages from FYTM % sheet"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    if 'FYTM %' not in wb.sheetnames:
        return None
    
    ws = wb['FYTM %']
    
    # Get the percentage row (appears to be row 4 based on formulas)
    percentages = []
    for col in range(2, 14):  # B through M (12 months)
        cell = ws.cell(4, col)
        if cell.value is not None:
            percentages.append(float(cell.value))
    
    return {
        'monthly_percentages': percentages,
        'total': sum(percentages)
    }

def extract_planned_order_logic(file_path):
    """Extract key columns from Planned order sheet"""
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    # Find the planned order sheet
    planned_sheet = None
    for name in wb.sheetnames:
        if 'planned' in name.lower() and 'order' in name.lower():
            planned_sheet = name
            break
    
    if not planned_sheet:
        return None
    
    ws = wb[planned_sheet]
    
    # Get headers
    headers = []
    for col in range(1, min(30, ws.max_column + 1)):
        cell = ws.cell(1, col)
        if cell.value:
            headers.append(str(cell.value))
    
    # Find key columns
    forecast_col = None
    soh_col = None
    stockout_col = None
    pending_po_col = None
    
    for i, h in enumerate(headers, 1):
        h_lower = h.lower()
        if 'forecast' in h_lower:
            forecast_col = i
        if 'soh' in h_lower or 'stock on hand' in h_lower:
            soh_col = i
        if 'stockout' in h_lower or 'stock out' in h_lower:
            stockout_col = i
        if 'pending' in h_lower and 'po' in h_lower:
            pending_po_col = i
    
    # Get sample formulas from key columns
    formulas = {}
    if stockout_col:
        for row in range(2, min(12, ws.max_row + 1)):
            cell = ws.cell(row, stockout_col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas[f'stockout_row_{row}'] = cell.value
                break
    
    if forecast_col:
        for row in range(2, min(12, ws.max_row + 1)):
            cell = ws.cell(row, forecast_col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas[f'forecast_row_{row}'] = cell.value
                break
    
    return {
        'sheet_name': planned_sheet,
        'headers': headers,
        'forecast_column': forecast_col,
        'soh_column': soh_col,
        'stockout_column': stockout_col,
        'pending_po_column': pending_po_col,
        'sample_formulas': formulas
    }

def extract_sales_history_structure(file_path):
    """Look at how sales history is structured"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Check IF MAR and similar sheets
    result = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in ['IF MAR', 'FYTM MAR', 'Forecast']:
            ws = wb[sheet_name]
            
            # Get first few rows to understand structure
            sample_rows = []
            for row_idx in range(1, min(6, ws.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(15, ws.max_column + 1)):
                    cell = ws.cell(row_idx, col_idx)
                    val = cell.value
                    if val:
                        row_data.append(str(val)[:50])
                sample_rows.append(row_data)
            
            result[sheet_name] = {
                'sample_rows': sample_rows,
                'dimensions': str(ws.dimensions)
            }
    
    return result

# Analyze all files
files = {
    'file_1': 'C:/Users/User/.openclaw/workspace/forecasting-1.xlsx',
    'file_2': 'C:/Users/User/.openclaw/workspace/forecasting-2.xlsx',
    'file_3': 'C:/Users/User/.openclaw/workspace/forecasting-3.xlsx'
}

results = {}
for key, f in files.items():
    try:
        results[key] = {
            'fytm_percentages': extract_fytm_percentages(f),
            'planned_order': extract_planned_order_logic(f),
            'sales_history': extract_sales_history_structure(f)
        }
    except Exception as e:
        print(f"Error analyzing {f}: {e}")
        results[key] = {'error': str(e)}

# Save results
with open('C:/Users/User/.openclaw/workspace/forecast-logic.json', 'w') as out:
    json.dump(results, out, indent=2)

# Print summary
print("="*80)
print("FORECASTING LOGIC ANALYSIS")
print("="*80)

for key, data in results.items():
    print(f"\n{key.upper()}:")
    
    if 'error' in data:
        print(f"  Error: {data['error']}")
        continue
    
    if data['fytm_percentages']:
        pct = data['fytm_percentages']
        print(f"\n  Monthly percentages (FYTM %):")
        print(f"    Values: {pct['monthly_percentages']}")
        print(f"    Total: {pct['total']}")
    
    if data['planned_order']:
        po = data['planned_order']
        print(f"\n  Planned Order sheet: '{po['sheet_name']}'")
        print(f"    Key columns:")
        if po['forecast_column']:
            print(f"      Forecast: Column {po['forecast_column']}")
        if po['soh_column']:
            print(f"      SOH: Column {po['soh_column']}")
        if po['stockout_column']:
            print(f"      Stockout: Column {po['stockout_column']}")
        if po['pending_po_column']:
            print(f"      Pending PO: Column {po['pending_po_column']}")
        
        if po['sample_formulas']:
            print(f"\n    Sample formulas:")
            for name, formula in po['sample_formulas'].items():
                print(f"      {name}: {formula[:150]}")
    
    if data['sales_history']:
        print(f"\n  Sales history sheets found: {list(data['sales_history'].keys())}")

print(f"\n\nDetailed results saved to: forecast-logic.json")
