import openpyxl

def get_key_formulas(file_path, sheet_name, columns_to_check):
    """Extract formulas from specific columns"""
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    if sheet_name not in wb.sheetnames:
        return f"Sheet '{sheet_name}' not found"
    
    ws = wb[sheet_name]
    
    formulas = {}
    
    # Check first 20 data rows
    for row_idx in range(2, min(22, ws.max_row + 1)):
        row_formulas = {}
        has_data = False
        
        for col_name, col_idx in columns_to_check.items():
            cell = ws.cell(row_idx, col_idx)
            if cell.value:
                has_data = True
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    row_formulas[col_name] = cell.value
                else:
                    row_formulas[col_name] = f"<value: {str(cell.value)[:50]}>"
        
        if has_data and row_formulas:
            formulas[f'row_{row_idx}'] = row_formulas
    
    return formulas

# Check File 2 (most complete structure)
print("="*80)
print("FILE 2: Planned order formulas")
print("="*80)

columns = {
    'FORECAST': 5,
    'SOH': 6,
    'Stockout_Month': 9,
    'PENDING_PO': 10,
    'Stockout_After_PO': 11,
    'QTY_TO_ORDER': 14
}

formulas = get_key_formulas(
    'C:/Users/User/.openclaw/workspace/forecasting-2.xlsx',
    'Planned order',
    columns
)

for row, data in list(formulas.items())[:5]:  # Show first 5 rows
    print(f"\n{row}:")
    for col, formula in data.items():
        print(f"  {col}: {formula}")

# Also check the Forecast sheet to see how monthly sales are calculated
print("\n" + "="*80)
print("FILE 2: Forecast sheet sample")
print("="*80)

wb = openpyxl.load_workbook('C:/Users/User/.openclaw/workspace/forecasting-2.xlsx', data_only=False)
ws = wb['Forecast']

# Get a sample item with formulas
for row_idx in range(6, min(16, ws.max_row + 1)):
    row_data = {}
    has_formula = False
    
    # Check first 15 columns
    for col_idx in range(1, 16):
        cell = ws.cell(row_idx, col_idx)
        if cell.value:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                row_data[f'col_{col_idx}'] = cell.value[:150]
                has_formula = True
            else:
                row_data[f'col_{col_idx}'] = f"<{str(cell.value)[:30]}>"
    
    if has_formula:
        print(f"\nRow {row_idx}:")
        for col, val in row_data.items():
            print(f"  {col}: {val}")
        break  # Just show one example

print("\n" + "="*80)
