import openpyxl
import pandas as pd
import json
from pathlib import Path

def analyze_excel(file_path):
    """Analyze Excel forecasting spreadsheet structure and formulas"""
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    results = {
        'filename': Path(file_path).name,
        'sheets': [],
        'key_patterns': {}
    }
    
    # Analyze each visible sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Skip if sheet is hidden
        if ws.sheet_state == 'hidden':
            continue
            
        sheet_info = {
            'name': sheet_name,
            'dimensions': str(ws.dimensions),
            'headers': [],
            'sample_formulas': [],
            'sample_data': []
        }
        
        # Get first row (likely headers)
        if ws.max_row > 0:
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))
            sheet_info['headers'] = headers[:20]  # First 20 headers
        
        # Look for formulas in first 50 rows, first 20 cols
        formulas_found = 0
        for row in range(1, min(51, ws.max_row + 1)):
            for col in range(1, min(21, ws.max_column + 1)):
                cell = ws.cell(row, col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    if formulas_found < 10:  # Sample first 10 formulas
                        sheet_info['sample_formulas'].append({
                            'cell': f'{cell.column_letter}{cell.row}',
                            'formula': cell.value[:200]  # Truncate long formulas
                        })
                        formulas_found += 1
        
        # Get sample data from first 10 rows
        for row_idx in range(2, min(12, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(11, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                val = cell.value
                if val and not (isinstance(val, str) and val.startswith('=')):
                    row_data.append(str(val)[:50])
                else:
                    row_data.append('')
            if any(row_data):
                sheet_info['sample_data'].append(row_data)
        
        results['sheets'].append(sheet_info)
    
    # Look for key patterns
    for sheet_info in results['sheets']:
        headers_lower = [h.lower() for h in sheet_info['headers']]
        
        # Check for sales history indicators
        if any('history' in h or 'sales' in h or 'sold' in h for h in headers_lower):
            results['key_patterns']['has_sales_history'] = sheet_info['name']
        
        # Check for seasonality/percentage indicators
        if any('%' in h or 'percent' in h or 'season' in h or 'month' in h for h in headers_lower):
            results['key_patterns']['has_seasonality'] = sheet_info['name']
        
        # Check for forecast/projection
        if any('forecast' in h or 'project' in h or 'expected' in h for h in headers_lower):
            results['key_patterns']['has_forecast'] = sheet_info['name']
        
        # Check for reorder/order quantity
        if any('order' in h or 'reorder' in h or 'qty' in h for h in headers_lower):
            results['key_patterns']['has_order_calc'] = sheet_info['name']
    
    return results

# Analyze all three spreadsheets
files = [
    'C:/Users/User/.openclaw/workspace/forecasting-1.xlsx',
    'C:/Users/User/.openclaw/workspace/forecasting-2.xlsx',
    'C:/Users/User/.openclaw/workspace/forecasting-3.xlsx'
]

all_results = []
for f in files:
    try:
        result = analyze_excel(f)
        all_results.append(result)
    except Exception as e:
        print(f"Error analyzing {f}: {e}")

# Save detailed results
with open('C:/Users/User/.openclaw/workspace/forecast-analysis.json', 'w') as out:
    json.dump(all_results, out, indent=2)

# Print summary
for result in all_results:
    print(f"\n{'='*60}")
    print(f"FILE: {result['filename']}")
    print(f"{'='*60}")
    print(f"\nVisible sheets: {len(result['sheets'])}")
    for sheet in result['sheets']:
        print(f"\n  Sheet: {sheet['name']}")
        print(f"  Headers ({len(sheet['headers'])}): {', '.join(sheet['headers'][:10])}")
        if sheet['sample_formulas']:
            print(f"  Sample formulas:")
            for f in sheet['sample_formulas'][:3]:
                print(f"    {f['cell']}: {f['formula']}")
    
    if result['key_patterns']:
        print(f"\n  Key patterns found:")
        for k, v in result['key_patterns'].items():
            print(f"    {k}: {v}")

print(f"\n\nDetailed results saved to: forecast-analysis.json")
